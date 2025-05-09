import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import EXCEL_FOLDER
import os

def log_success(message):
    print(f"[+] {message}")

def log_failure(message):
    print(f"[X] {message}")

def log_info(message):
    print(f"[!] {message}")

def create_template(data, filename):
    """Membuat template Excel dengan data yang diberikan"""
    # Pastikan folder output ada
    if not os.path.exists(EXCEL_FOLDER):
        os.makedirs(EXCEL_FOLDER)
        log_success(f"Created directory: {EXCEL_FOLDER}")
        
    # Gabungkan folder dan nama file
    full_filepath = os.path.join(EXCEL_FOLDER, filename)

    # Buat DataFrame
    df = pd.DataFrame(data)
    
    # Simpan ke Excel di folder yang ditentukan
    try:
        df.to_excel(full_filepath, index=False)
        # log_success(f"DataFrame saved to {full_filepath}")
    except Exception as e:
        log_failure(f"Failed to save DataFrame to {full_filepath}: {e}")
        return

    # Buka workbook untuk styling
    try:
        wb = openpyxl.load_workbook(full_filepath)
        # log_success(f"Workbook {full_filepath} opened successfully")
    except Exception as e:
        log_failure(f"Error: Tidak dapat membuka file {full_filepath} - {e}")
        return
        
    ws = wb.active
    if ws is None:
        log_failure(f"Error: Tidak dapat mengakses worksheet dari file {full_filepath}")
        return
    
    # Styling
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Styling header
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = alignment
    
    # Styling data
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = alignment
    
    # Set column widths
    column_widths = {
        'A': 30,  # teknik
        'B': 40,  # deskripsi 
        'C': 40,  # gejala 
        'D': 80,  # solusi 
        'E': 30,  # tools (sebelumnya F)
        'F': 40,  # referensi (sebelumnya G)
        'G': 15,  # prioritas (sebelumnya H)
        'H': 40,  # masalah (sebelumnya I)
        'I': 40   # tanda_serangan (sebelumnya J)
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Simpan perubahan
    try:
        wb.save(full_filepath)
        log_success(f"Template {full_filepath} berhasil dibuat!")
    except Exception as e:
        log_failure(f"Error saat menyimpan file {full_filepath}: {str(e)}")

# Data untuk Network Security
network_data = {
    'teknik': [
        'Man-in-the-Middle (MITM) Attack',
        'ARP Spoofing',
        'DNS Spoofing',
        'WiFi Deauthentication',
        'WEP/WPA Cracking',
        'Evil Twin Attack',
        'Rogue Access Point',
        'Packet Sniffing',
        'MAC Spoofing',
        'DHCP Starvation'
    ],
    'deskripsi': [
        'Serangan yang memungkinkan penyerang untuk menyadap dan memodifikasi komunikasi antara dua pihak',
        'Serangan yang memanipulasi tabel ARP untuk mengarahkan traffic ke perangkat penyerang',
        'Serangan yang memanipulasi DNS untuk mengarahkan traffic ke server yang salah',
        'Serangan yang memutuskan koneksi client dari jaringan WiFi',
        'Serangan yang memecahkan enkripsi WEP/WPA untuk mendapatkan akses ke jaringan',
        'Serangan yang membuat titik akses WiFi palsu yang mirip dengan yang asli',
        'Serangan yang memasang titik akses WiFi yang tidak sah dalam jaringan',
        'Serangan yang menangkap dan menganalisis paket data dalam jaringan',
        'Serangan yang mengubah alamat MAC untuk menyembunyikan identitas',
        'Serangan yang menghabiskan pool alamat IP DHCP'
    ],
    'gejala': [
        'Traffic redirection|Data interception|Session hijacking',
        'ARP cache poisoning|Traffic redirection|Session hijacking',
        'DNS cache poisoning|Traffic redirection|Phishing',
        'Disconnection of clients|Denial of service|Session hijacking',
        'Cracking of encryption keys|Unauthorized access|Data theft',
        'Traffic redirection|Phishing|Credential theft',
        'Traffic redirection|Man-in-the-middle|Data theft',
        'Data interception|Credential theft|Session hijacking',
        'Identity spoofing|Unauthorized access|Bypassing MAC filtering',
        'Exhaustion of DHCP pool|Denial of service|Unauthorized IP assignment'
    ],
    'solusi': [
        'Gunakan HTTPS di seluruh aplikasi dan situs web untuk mengenkripsi komunikasi data. | Dapatkan sertifikat SSL/TLS yang valid dari penyedia terpercaya seperti Let\'s Encrypt (https://letsencrypt.org/). | Pastikan juga server dikonfigurasi dengan HTTP Strict Transport Security (HSTS) untuk mencegah downgrade attack. | Untuk memperkuat perlindungan, implementasikan certificate pinning di aplikasi klien agar hanya sertifikat dari server yang sah yang diterima. | Selain itu, validasi sertifikat dengan benar di sisi klien dan server agar tidak rentan terhadap serangan MITM.',

        'Untuk mencegah serangan ARP spoofing, terapkan Dynamic ARP Inspection (DAI) pada switch yang mendukung fitur ini. | Pertimbangkan untuk menggunakan static ARP entries pada perangkat yang krusial. | Monitoring ARP traffic secara aktif menggunakan tools seperti Arpwatch (https://linux.die.net/man/8/arpwatch) juga dapat membantu mendeteksi perubahan mendadak atau aktivitas mencurigakan dalam jaringan.',

        'Gunakan sistem DNS yang aman untuk melindungi dari serangan seperti DNS spoofing atau poisoning. | Implementasikan DNSSEC (Domain Name System Security Extensions) untuk memastikan integritas data DNS. | Gunakan DNS over HTTPS (DoH) atau DNS over TLS (DoT) untuk mengenkripsi permintaan DNS. | Penyedia DNS seperti Cloudflare (https://1.1.1.1), Google DNS (https://developers.google.com/speed/public-dns), atau NextDNS menyediakan fitur filtering dan keamanan tambahan.',

        'Lindungi jaringan nirkabel dari serangan deauthentication dengan menggunakan standar keamanan terbaru seperti WPA3. | Aktifkan client isolation pada access point untuk mencegah komunikasi langsung antar klien. | Monitoring deauth frames menggunakan tools seperti Kismet (https://www.kismetwireless.net/) dapat membantu mendeteksi serangan dini terhadap Wi-Fi.',

        'Selalu gunakan metode otentikasi yang kuat di jaringan Wi-Fi. | Hindari WEP dan WPA karena sudah usang dan rentan. | Pilih WPA3 dengan passphrase yang kuat dan kompleks. | Pertimbangkan juga untuk mengimplementasikan otentikasi berbasis sertifikat melalui protokol 802.1X dan RADIUS server untuk keamanan tingkat enterprise.',

        'Untuk mencegah serangan Evil Twin, validasi SSID secara manual di perangkat klien. | Implementasikan certificate pinning di aplikasi mobile agar koneksi hanya terjadi ke server yang sah. | Gunakan tools seperti Kismet untuk melakukan monitoring sinyal Wi-Fi dan mendeteksi access point palsu yang mencoba meniru SSID asli.',

        'Deteksi access point berbahaya seperti rogue AP bisa dilakukan dengan implementasi Wireless Intrusion Detection System (WIDS). | Monitoring SSID yang tidak dikenal dapat membantu mendeteksi AP mencurigakan. | Tools seperti Aircrack-ng atau Kismet bisa digunakan untuk memantau spektrum wireless dan mendeteksi AP mencurigakan yang muncul di jaringan.',

        'Untuk melindungi dari sniffing data, enkripsi perlu diterapkan di semua lapisan komunikasi (end-to-end). | Lakukan segmentasi jaringan untuk memisahkan trafik sensitif dari publik. | Lalu lintas dapat dianalisis menggunakan Wireshark (https://www.wireshark.org/) untuk mendeteksi apakah ada penyadapan atau traffic abnormal yang terjadi.',

        'Atasi spoofing MAC address dengan mengaktifkan fitur MAC filtering dan port security pada switch. | Batasi jumlah MAC address per port serta log semua perubahan. | Gunakan tools seperti Nmap (https://nmap.org/) untuk mendeteksi perangkat dengan MAC address duplikat atau mencurigakan dalam jaringan.',

        'Lindungi DHCP server dari serangan starvation dan spoofing dengan mengaktifkan fitur DHCP snooping di switch. | Terapkan rate limiting untuk membatasi permintaan DHCP dari satu host. | Monitoring dapat dilakukan dengan sistem logging dan tools seperti Yersinia untuk uji coba keamanan protokol layer 2.'
    ],
    'tools': [
        'Wireshark|Ettercap|SSLstrip|mitmproxy',
        'Ettercap|Arpwatch|Cain & Abel|Wireshark',
        'DNS spoofing tools|DNS monitoring|DNS security tools',
        'Aircrack-ng|Kismet|Wireshark|Deauthentication tools',
        'Aircrack-ng|Hashcat|John the Ripper|WPA cracking tools',
        'Kismet|Wireshark|Evil twin tools|Monitoring tools',
        'Kismet|Wireshark|Rogue AP detection|Monitoring tools',
        'Wireshark|tcpdump|NetworkMiner|Packet analysis tools',
        'Nmap|MAC changer|Network monitoring|Security tools',
        'DHCP monitoring|Network tools|Security tools'
    ],
    'referensi': [
        'https://www.cisco.com/c/en/us/solutions/security/man-in-the-middle-attack.html|https://www.owasp.org/index.php/Man-in-the-middle_attack',
        'https://www.networkworld.com/article/3278446/how-to-prevent-arp-spoofing-attacks.html|https://www.owasp.org/index.php/ARP_spoofing',
        'https://www.cloudflare.com/learning/dns/dns-security/|https://www.owasp.org/index.php/DNS_spoofing',
        'https://www.wi-fi.org/discover-wi-fi/security|https://www.owasp.org/index.php/Wi-Fi_security',
        'https://www.wi-fi.org/discover-wi-fi/security|https://www.owasp.org/index.php/WEP_cracking',
        'https://www.kaspersky.com/resource-center/definitions/evil-twin|https://www.owasp.org/index.php/Evil_twin_attack',
        'https://www.cisco.com/c/en/us/solutions/security/wireless-security.html|https://www.owasp.org/index.php/Rogue_access_point',
        'https://www.wireshark.org/docs/wsug_html_chunked/|https://www.owasp.org/index.php/Packet_sniffing',
        'https://nmap.org/docs.html|https://www.owasp.org/index.php/MAC_spoofing',
        'https://www.cisco.com/c/en/us/solutions/security/dhcp-snooping.html|https://www.owasp.org/index.php/DHCP_starvation'
    ],
    'prioritas': [
        'Kritis',
        'Tinggi',
        'Tinggi',
        'Tinggi',
        'Kritis',
        'Tinggi',
        'Tinggi',
        'Tinggi',
        'Sedang',
        'Sedang'
    ],
    'masalah': [
        'Vulnerability pada komunikasi jaringan yang tidak terenkripsi|Lack of certificate validation|Insufficient traffic monitoring',
        'Vulnerability pada protokol ARP yang tidak memiliki mekanisme autentikasi|Lack of ARP inspection|Insufficient monitoring',
        'Vulnerability pada protokol DNS yang tidak memiliki mekanisme autentikasi|Lack of DNSSEC|Insufficient DNS security',
        'Vulnerability pada protokol WiFi yang memungkinkan deauthentication|Lack of client isolation|Insufficient monitoring',
        'Vulnerability pada protokol enkripsi WiFi yang lemah|Weak password policies|Legacy protocol support',
        'Vulnerability pada autentikasi WiFi yang memungkinkan evil twin|Lack of certificate validation|Insufficient monitoring',
        'Vulnerability pada infrastruktur WiFi yang memungkinkan rogue AP|Lack of AP validation|Insufficient monitoring',
        'Vulnerability pada protokol jaringan yang tidak terenkripsi|Lack of network segmentation|Insufficient monitoring',
        'Vulnerability pada protokol MAC yang tidak memiliki mekanisme autentikasi|Lack of MAC filtering|Insufficient monitoring',
        'Vulnerability pada protokol DHCP yang tidak memiliki mekanisme autentikasi|Lack of DHCP snooping|Insufficient monitoring'
    ],
    'tanda_serangan': [
        'Traffic redirection|Data interception|Session hijacking|Certificate errors',
        'ARP cache poisoning|Traffic redirection|Session hijacking|Network anomalies',
        'DNS cache poisoning|Traffic redirection|Phishing|DNS errors',
        'Disconnection of clients|Denial of service|Session hijacking|Network instability',
        'Cracking of encryption keys|Unauthorized access|Data theft|Network anomalies',
        'Traffic redirection|Phishing|Credential theft|Network anomalies',
        'Traffic redirection|Man-in-the-middle|Data theft|Network anomalies',
        'Data interception|Credential theft|Session hijacking|Network anomalies',
        'Identity spoofing|Unauthorized access|Bypassing MAC filtering|Network anomalies',
        'Exhaustion of DHCP pool|Denial of service|Unauthorized IP assignment|Network instability'
    ]
}

# Data untuk Web Security
web_data = {
    'teknik': [
        'SQL Injection',
        'Cross-Site Scripting (XSS)',
        'Cross-Site Request Forgery (CSRF)',
        'File Upload Vulnerability',
        'Insecure Direct Object References',
        'Security Misconfiguration',
        'Broken Authentication',
        'Sensitive Data Exposure',
        'Using Components with Known Vulnerabilities',
        'Insufficient Logging & Monitoring'
    ],
    'deskripsi': [
        'Serangan yang memanfaatkan celah keamanan pada query database',
        'Serangan yang menyisipkan script berbahaya ke halaman web',
        'Serangan yang memaksa pengguna melakukan aksi tanpa sepengetahuan mereka',
        'Serangan yang memanfaatkan fitur upload file untuk mengeksekusi kode berbahaya',
        'Serangan yang mengakses objek secara langsung tanpa validasi',
        'Serangan yang memanfaatkan konfigurasi keamanan yang tidak tepat',
        'Serangan yang memanfaatkan celah pada sistem autentikasi',
        'Serangan yang mengekspos data sensitif',
        'Serangan yang memanfaatkan komponen dengan celah keamanan',
        'Serangan yang memanfaatkan kurangnya logging dan monitoring'
    ],
    'gejala': [
        'Data manipulation|Unauthorized access|Database errors',
        'Script injection|Session hijacking|Data theft',
        'Unauthorized actions|Session manipulation|Data modification',
        'Malware execution|Server compromise|Data theft',
        'Unauthorized access|Data exposure|Privilege escalation',
        'Security bypass|Unauthorized access|System compromise',
        'Account takeover|Session hijacking|Unauthorized access',
        'Data exposure|Privacy violation|Compliance issues',
        'System compromise|Data theft|Service disruption',
        'Security incidents|Data breaches|System compromise'
    ],
    'solusi': [
        'Gunakan Parameterized Queries (Prepared Statements): Ini cara paling efektif mencegah SQLi, pisahkan kode SQL dari data input. | Validasi Input Secara Ketat: Terapkan whitelist untuk format data yang diharapkan (angka, email, dll.) dan tolak input yang tidak sesuai. | Gunakan Object-Relational Mapper (ORM): Banyak ORM menangani parameterisasi secara otomatis (misal SQLAlchemy, Django ORM). | Prinsip Least Privilege: Berikan hak akses database seminimal mungkin untuk akun aplikasi web. | Terapkan Web Application Firewall (WAF): Gunakan WAF (misal ModSecurity [https://modsecurity.org/]) untuk memfilter request berbahaya.',

        'Sanitasi Input dengan Benar: Hapus karakter atau tag berbahaya sebelum diproses. | Gunakan Output Encoding: Escape karakter khusus saat output ditampilkan ke browser (misalnya `&`, `<`, `>`). | Implementasi Content Security Policy (CSP): Batasi resource yang bisa dimuat dari luar domain. | Gunakan Filter XSS: Framework modern biasanya memiliki proteksi XSS bawaan yang perlu diaktifkan atau dikonfigurasi.',

        'Gunakan CSRF Token pada Semua Form yang Mengubah Data: Token unik per sesi mencegah request palsu. | Terapkan Cookie dengan Atribut SameSite: Gunakan `SameSite=Strict` atau `Lax` agar cookie tidak dikirim ke domain berbeda. | Validasi Origin dan Referer: Pastikan request datang dari domain yang sah. | Validasi Semua Request Penting di Server: Jangan hanya mengandalkan sisi klien.',

        'Validasi Tipe File: Izinkan hanya jenis file tertentu (misalnya hanya .pdf atau .jpg). | Pindai File Terupload dari Malware: Gunakan antivirus engine seperti ClamAV atau layanan scanning eksternal. | Simpan File dengan Aman: Jangan taruh di direktori publik secara langsung, dan ubah nama file agar tidak bisa diakses langsung. | Atur Akses File: Hanya user yang berhak boleh mengakses file tersebut, gunakan autentikasi dan izin akses yang tepat.',

        'Terapkan Kontrol Akses Ketat: Pastikan hanya user yang berwenang dapat mengakses atau memodifikasi objek. | Validasi Input yang Digunakan untuk Akses Objek: Hindari penggunaan ID yang mudah ditebak. | Validasi Kepemilikan Objek di Server: Jangan percaya data dari klien. | Implementasi Middleware atau Guard: Cek hak akses secara sistematis di semua endpoint.',

        'Pasang Security Headers: Gunakan header seperti `X-Content-Type-Options`, `X-Frame-Options`, `Strict-Transport-Security`. | Konfigurasi Aplikasi dengan Aman: Matikan fitur default yang tidak digunakan, hindari info leakage (seperti error stack trace). | Update Rutin Semua Komponen: Termasuk framework, library, dan dependensi pihak ketiga. | Lakukan Pengujian Keamanan Berkala: Gunakan automated scanner atau manual penetration testing.',

        'Gunakan Otentikasi yang Kuat: Gunakan metode seperti OAuth2 atau JWT dengan mekanisme validasi yang benar. | Kelola Sesi Secara Aman: Gunakan ID sesi yang tidak bisa ditebak, simpan di cookie dengan `HttpOnly` dan `Secure`. | Terapkan Kebijakan Password yang Baik: Minimal panjang, kompleksitas, dan rotasi berkala. | Tambahkan 2FA: Gunakan autentikasi dua faktor (misalnya TOTP atau OTP via SMS/email) untuk lapisan keamanan tambahan.',

        'Enkripsi Data Sensitif: Gunakan algoritma modern seperti AES-256 untuk penyimpanan data rahasia. | Simpan Data secara Aman: Hindari penyimpanan langsung di local storage untuk data penting. | Atur Akses Data dengan Ketat: Implementasikan role-based access control (RBAC). | Minimalkan Data yang Dikumpulkan: Ambil hanya data yang benar-benar dibutuhkan.',

        'Update Sistem Secara Berkala: Pastikan sistem operasi dan semua aplikasi mendapatkan patch terbaru. | Lakukan Pemindaian Kerentanan: Gunakan tools seperti OpenVAS atau Nessus secara berkala. | Kelola Inventaris Komponen: Buat daftar semua library dan software yang digunakan. | Terapkan Manajemen Patch: Buat prosedur untuk uji coba dan deploy patch dengan aman.',

        'Aktifkan Logging yang Komprehensif: Catat semua aktivitas penting, login, error, dan request mencurigakan. | Monitoring Real-Time: Gunakan sistem monitoring seperti ELK Stack, Grafana, atau SIEM. | Sistem Notifikasi dan Alert: Kirim peringatan bila ada aktivitas anomali atau percobaan serangan. | Siapkan Prosedur Tanggap Insiden: Tim harus tahu langkah-langkah yang dilakukan saat insiden terjadi.'
    ],
    'tools': [
        'SQLmap|Acunetix|Burp Suite|OWASP ZAP',
        'XSS Scanner|Burp Suite|OWASP ZAP|Acunetix',
        'CSRF Scanner|Burp Suite|OWASP ZAP|Acunetix',
        'File upload scanner|Malware scanner|Burp Suite|OWASP ZAP',
        'IDOR scanner|Burp Suite|OWASP ZAP|Acunetix',
        'Configuration scanner|Security scanner|Burp Suite|OWASP ZAP',
        'Auth scanner|Session analyzer|Burp Suite|OWASP ZAP',
        'Data exposure scanner|Encryption checker|Burp Suite|OWASP ZAP',
        'Vulnerability scanner|Component checker|Burp Suite|OWASP ZAP',
        'Log analyzer|Monitoring tools|SIEM|Security tools'
    ],
    'referensi': [
        'https://www.owasp.org/index.php/SQL_Injection|https://www.w3schools.com/sql/sql_injection.asp',
        'https://www.owasp.org/index.php/Cross-site_Scripting_(XSS)|https://www.w3schools.com/js/js_xss.asp',
        'https://www.owasp.org/index.php/Cross-Site_Request_Forgery_(CSRF)|https://www.w3schools.com/php/php_csrf.asp',
        'https://www.owasp.org/index.php/Unrestricted_File_Upload|https://www.w3schools.com/php/php_file_upload.asp',
        'https://www.owasp.org/index.php/Insecure_Direct_Object_References|https://www.w3schools.com/php/php_security.asp',
        'https://www.owasp.org/index.php/Security_Misconfiguration|https://www.w3schools.com/php/php_security.asp',
        'https://www.owasp.org/index.php/Broken_Authentication|https://www.w3schools.com/php/php_security.asp',
        'https://www.owasp.org/index.php/Sensitive_Data_Exposure|https://www.w3schools.com/php/php_security.asp',
        'https://www.owasp.org/index.php/Using_Components_with_Known_Vulnerabilities|https://www.w3schools.com/php/php_security.asp',
        'https://www.owasp.org/index.php/Insufficient_Logging_%26_Monitoring|https://www.w3schools.com/php/php_security.asp'
    ],
    'prioritas': [
        'Kritis',
        'Kritis',
        'Tinggi',
        'Tinggi',
        'Tinggi',
        'Tinggi',
        'Kritis',
        'Kritis',
        'Tinggi',
        'Tinggi'
    ],
    'masalah': [
        'Vulnerability pada query database|Lack of input validation|Insufficient output encoding',
        'Vulnerability pada rendering HTML|Lack of input sanitization|Insufficient output encoding',
        'Vulnerability pada request handling|Lack of CSRF protection|Insufficient validation',
        'Vulnerability pada file handling|Lack of file validation|Insufficient security',
        'Vulnerability pada object access|Lack of access control|Insufficient validation',
        'Vulnerability pada configuration|Lack of security headers|Insufficient security',
        'Vulnerability pada authentication|Lack of session management|Insufficient security',
        'Vulnerability pada data handling|Lack of encryption|Insufficient security',
        'Vulnerability pada components|Lack of updates|Insufficient security',
        'Vulnerability pada logging|Lack of monitoring|Insufficient security'
    ],
    'tanda_serangan': [
        'Database errors|Data manipulation|Unauthorized access|SQL errors',
        'Script execution|Session hijacking|Data theft|XSS alerts',
        'Unauthorized actions|Session manipulation|Data modification|CSRF alerts',
        'Malware execution|Server compromise|Data theft|File alerts',
        'Unauthorized access|Data exposure|Privilege escalation|IDOR alerts',
        'Security bypass|Unauthorized access|System compromise|Config alerts',
        'Account takeover|Session hijacking|Unauthorized access|Auth alerts',
        'Data exposure|Privacy violation|Compliance issues|Data alerts',
        'System compromise|Data theft|Service disruption|Vuln alerts',
        'Security incidents|Data breaches|System compromise|Log alerts'
    ]
}

# Data untuk Mobile Security
mobile_data = {
    'teknik': [
        'Reverse Engineering',
        'Man-in-the-Middle Attack',
        'Malware Injection',
        'Insecure Data Storage',
        'Weak Encryption',
        'Insecure Communication',
        'Code Tampering',
        'Root/Jailbreak Detection Bypass',
        'Insecure Authentication',
        'Insecure Authorization'
    ],
    'deskripsi': [
        'Serangan yang menganalisis dan memodifikasi aplikasi mobile',
        'Serangan yang menyadap komunikasi antara aplikasi dan server',
        'Serangan yang menyisipkan malware ke aplikasi mobile',
        'Serangan yang mengekspos data yang disimpan secara tidak aman',
        'Serangan yang memanfaatkan enkripsi yang lemah',
        'Serangan yang menyadap komunikasi yang tidak aman',
        'Serangan yang memodifikasi kode aplikasi',
        'Serangan yang memanfaatkan root/jailbreak untuk bypass keamanan',
        'Serangan yang memanfaatkan autentikasi yang lemah',
        'Serangan yang memanfaatkan otorisasi yang lemah'
    ],
    'gejala': [
        'Code analysis|Binary modification|App tampering',
        'Traffic interception|Data theft|Session hijacking',
        'Malware presence|System compromise|Data theft',
        'Data exposure|Privacy violation|Compliance issues',
        'Encryption bypass|Data exposure|Privacy violation',
        'Traffic interception|Data theft|Privacy violation',
        'Code modification|App tampering|Security bypass',
        'Security bypass|Privilege escalation|System compromise',
        'Account takeover|Session hijacking|Unauthorized access',
        'Privilege escalation|Unauthorized access|Data exposure'
    ],
    'solusi': [
        'Code obfuscation|Anti-tampering|Integrity checks|Security testing',
        'Certificate pinning|Encryption|Secure communication|Traffic monitoring',
        'Malware scanning|Code signing|Integrity checks|Security testing',
        'Secure storage|Encryption|Access control|Data minimization',
        'Strong encryption|Key management|Secure storage|Regular updates',
        'SSL/TLS|Certificate validation|Secure communication|Traffic monitoring',
        'Code signing|Integrity checks|Anti-tampering|Security testing',
        'Root detection|Jailbreak detection|Integrity checks|Security testing',
        'Strong authentication|Session management|2FA|Security testing',
        'Access control|Authorization checks|Security testing|Regular updates'
    ],
    'tools': [
        'IDA Pro|Hopper|APKTool|JEB Decompiler',
        'Burp Suite|Fiddler|Wireshark|SSLstrip',
        'Malware scanner|Code analyzer|Security tools|Monitoring',
        'Storage analyzer|Encryption checker|Security tools|Monitoring',
        'Encryption analyzer|Key checker|Security tools|Monitoring',
        'Network analyzer|SSL checker|Security tools|Monitoring',
        'Code analyzer|Integrity checker|Security tools|Monitoring',
        'Root checker|Jailbreak detector|Security tools|Monitoring',
        'Auth analyzer|Session checker|Security tools|Monitoring',
        'Auth analyzer|Access checker|Security tools|Monitoring'
    ],
    'referensi': [
        'https://www.owasp.org/index.php/Mobile_Top_10_2016-M1-Weak_Server_Side_Controls|https://www.owasp.org/index.php/Reverse_Engineering',
        'https://www.owasp.org/index.php/Mobile_Top_10_2016-M3-Insecure_Communication|https://www.owasp.org/index.php/Mobile_Top_10_2016-M3',
        'https://www.owasp.org/index.php/Mobile_Top_10_2016-M4-Insecure_Authentication|https://www.owasp.org/index.php/Mobile_Top_10_2016-M4',
        'https://www.owasp.org/index.php/Mobile_Top_10_2016-M5-Insufficient_Cryptography|https://www.owasp.org/index.php/Mobile_Top_10_2016-M5',
        'https://www.owasp.org/index.php/Mobile_Top_10_2016-M6-Insecure_Authorization|https://www.owasp.org/index.php/Mobile_Top_10_2016-M6',
        'https://www.owasp.org/index.php/Mobile_Top_10_2016-M7-Client_Code_Quality|https://www.owasp.org/index.php/Mobile_Top_10_2016-M7',
        'https://www.owasp.org/index.php/Mobile_Top_10_2016-M8-Code_Tampering|https://www.owasp.org/index.php/Mobile_Top_10_2016-M8',
        'https://www.owasp.org/index.php/Mobile_Top_10_2016-M9-Reverse_Engineering|https://www.owasp.org/index.php/Mobile_Top_10_2016-M9',
        'https://www.owasp.org/index.php/Mobile_Top_10_2016-M10-Extraneous_Functionality|https://www.owasp.org/index.php/Mobile_Top_10_2016-M10',
        'https://www.owasp.org/index.php/Mobile_Top_10_2016-M2-Insufficient_Transport_Layer_Protection|https://www.owasp.org/index.php/Mobile_Top_10_2016-M2'
    ],
    'prioritas': [
        'Kritis',
        'Kritis',
        'Kritis',
        'Tinggi',
        'Tinggi',
        'Tinggi',
        'Tinggi',
        'Tinggi',
        'Tinggi',
        'Tinggi'
    ],
    'masalah': [
        'Vulnerability pada kode aplikasi|Lack of code protection|Insufficient security',
        'Vulnerability pada komunikasi|Lack of encryption|Insufficient security',
        'Vulnerability pada aplikasi|Lack of malware protection|Insufficient security',
        'Vulnerability pada penyimpanan|Lack of encryption|Insufficient security',
        'Vulnerability pada enkripsi|Lack of key management|Insufficient security',
        'Vulnerability pada komunikasi|Lack of SSL/TLS|Insufficient security',
        'Vulnerability pada kode|Lack of integrity checks|Insufficient security',
        'Vulnerability pada sistem|Lack of root detection|Insufficient security',
        'Vulnerability pada autentikasi|Lack of session management|Insufficient security',
        'Vulnerability pada otorisasi|Lack of access control|Insufficient security'
    ],
    'tanda_serangan': [
        'Code modification|App tampering|Security bypass|Integrity alerts',
        'Traffic interception|Data theft|Privacy violation|Comm alerts',
        'Malware presence|System compromise|Data theft|Malware alerts',
        'Data exposure|Privacy violation|Compliance issues|Storage alerts',
        'Encryption bypass|Data exposure|Privacy violation|Crypto alerts',
        'Traffic interception|Data theft|Privacy violation|Comm alerts',
        'Code modification|App tampering|Security bypass|Code alerts',
        'Security bypass|Privilege escalation|System compromise|Root alerts',
        'Account takeover|Session hijacking|Unauthorized access|Auth alerts',
        'Privilege escalation|Unauthorized access|Data exposure|Auth alerts'
    ]
}

# Data untuk IoT Security
iot_data = {
    'teknik': [
        'Default Credentials',
        'Firmware Vulnerabilities',
        'Insecure Communication',
        'Physical Tampering',
        'Denial of Service',
        'Data Leakage',
        'Insecure Update Mechanism',
        'Weak Encryption',
        'Insecure Network Services',
        'Privilege Escalation'
    ],
    'deskripsi': [
        'Serangan yang memanfaatkan kredensial default pada perangkat IoT',
        'Serangan yang memanfaatkan celah keamanan pada firmware',
        'Serangan yang menyadap komunikasi yang tidak aman',
        'Serangan yang memanipulasi perangkat secara fisik',
        'Serangan yang membuat perangkat tidak dapat diakses',
        'Serangan yang mengekspos data sensitif',
        'Serangan yang memanfaatkan mekanisme update yang tidak aman',
        'Serangan yang memanfaatkan enkripsi yang lemah',
        'Serangan yang memanfaatkan layanan jaringan yang tidak aman',
        'Serangan yang meningkatkan hak akses pada perangkat'
    ],
    'gejala': [
        'Unauthorized access|Account takeover|System compromise',
        'System compromise|Data theft|Service disruption',
        'Data interception|Privacy violation|Service disruption',
        'Physical damage|System compromise|Data theft',
        'Service disruption|System unavailability|Resource exhaustion',
        'Data exposure|Privacy violation|Compliance issues',
        'System compromise|Malware injection|Service disruption',
        'Data exposure|Privacy violation|Service disruption',
        'Unauthorized access|Service disruption|System compromise',
        'Privilege escalation|Unauthorized access|System compromise'
    ],
    'solusi': [
        'Change default credentials|Strong passwords|Regular updates|Access control',
        'Regular firmware updates|Vulnerability scanning|Security testing|Patch management',
        'Encryption|Secure communication|Certificate validation|Traffic monitoring',
        'Physical security|Tamper detection|Secure storage|Access control',
        'DoS protection|Rate limiting|Traffic monitoring|Resource management',
        'Data encryption|Secure storage|Access control|Data minimization',
        'Secure update mechanism|Signature validation|Rollback protection|Update monitoring',
        'Strong encryption|Key management|Secure storage|Regular updates',
        'Service hardening|Access control|Network segmentation|Security monitoring',
        'Access control|Privilege management|Security monitoring|Regular updates'
    ],
    'tools': [
        'Credential scanner|Password tester|Security tools|Monitoring',
        'Firmware analyzer|Vulnerability scanner|Security tools|Monitoring',
        'Network analyzer|SSL checker|Security tools|Monitoring',
        'Physical security tools|Tamper detector|Security tools|Monitoring',
        'DoS tester|Network analyzer|Security tools|Monitoring',
        'Data analyzer|Encryption checker|Security tools|Monitoring',
        'Update analyzer|Signature checker|Security tools|Monitoring',
        'Encryption analyzer|Key checker|Security tools|Monitoring',
        'Service analyzer|Network scanner|Security tools|Monitoring',
        'Access analyzer|Privilege checker|Security tools|Monitoring'
    ],
    'referensi': [
        'https://www.owasp.org/index.php/IoT_Default_Credentials|https://www.owasp.org/index.php/IoT_Security',
        'https://www.owasp.org/index.php/IoT_Firmware_Vulnerabilities|https://www.owasp.org/index.php/IoT_Security',
        'https://www.owasp.org/index.php/IoT_Insecure_Communication|https://www.owasp.org/index.php/IoT_Security',
        'https://www.owasp.org/index.php/IoT_Physical_Tampering|https://www.owasp.org/index.php/IoT_Security',
        'https://www.owasp.org/index.php/IoT_Denial_of_Service|https://www.owasp.org/index.php/IoT_Security',
        'https://www.owasp.org/index.php/IoT_Data_Leakage|https://www.owasp.org/index.php/IoT_Security',
        'https://www.owasp.org/index.php/IoT_Insecure_Update_Mechanism|https://www.owasp.org/index.php/IoT_Security',
        'https://www.owasp.org/index.php/IoT_Weak_Encryption|https://www.owasp.org/index.php/IoT_Security',
        'https://www.owasp.org/index.php/IoT_Insecure_Network_Services|https://www.owasp.org/index.php/IoT_Security',
        'https://www.owasp.org/index.php/IoT_Privilege_Escalation|https://www.owasp.org/index.php/IoT_Security'
    ],
    'prioritas': [
        'Kritis',
        'Kritis',
        'Kritis',
        'Tinggi',
        'Tinggi',
        'Tinggi',
        'Tinggi',
        'Tinggi',
        'Tinggi',
        'Tinggi'
    ],
    'masalah': [
        'Vulnerability pada kredensial|Lack of password policy|Insufficient security',
        'Vulnerability pada firmware|Lack of updates|Insufficient security',
        'Vulnerability pada komunikasi|Lack of encryption|Insufficient security',
        'Vulnerability pada fisik|Lack of physical security|Insufficient security',
        'Vulnerability pada layanan|Lack of DoS protection|Insufficient security',
        'Vulnerability pada data|Lack of encryption|Insufficient security',
        'Vulnerability pada update|Lack of security|Insufficient security',
        'Vulnerability pada enkripsi|Lack of key management|Insufficient security',
        'Vulnerability pada layanan|Lack of hardening|Insufficient security',
        'Vulnerability pada akses|Lack of access control|Insufficient security'
    ],
    'tanda_serangan': [
        'Unauthorized access|Account takeover|System compromise|Auth alerts',
        'System compromise|Data theft|Service disruption|Firmware alerts',
        'Data interception|Privacy violation|Service disruption|Comm alerts',
        'Physical damage|System compromise|Data theft|Physical alerts',
        'Service disruption|System unavailability|Resource exhaustion|DoS alerts',
        'Data exposure|Privacy violation|Compliance issues|Data alerts',
        'System compromise|Malware injection|Service disruption|Update alerts',
        'Data exposure|Privacy violation|Service disruption|Crypto alerts',
        'Unauthorized access|Service disruption|System compromise|Service alerts',
        'Privilege escalation|Unauthorized access|System compromise|Access alerts'
    ]
}

# Definisikan fungsi utama untuk membuat semua template
def create_all_templates():
    create_template(network_data, 'network_security.xlsx')
    # create_template(web_data, 'web_security.xlsx')
    # create_template(mobile_data, 'mobile_security.xlsx')
    # create_template(iot_data, 'iot_security.xlsx')
    # print(f"\nSemua template Excel berhasil dibuat di folder {EXCEL_FOLDER}")

# Jalankan pembuatan template jika script dijalankan langsung (opsional, tapi baik ada)
if __name__ == "__main__":
    create_all_templates() 