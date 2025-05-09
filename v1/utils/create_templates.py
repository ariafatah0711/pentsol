import os
import json
import pandas as pd
from config import DATA_FILE, EXCEL_TEMPLATE
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import EXCEL_FOLDER

def log_success(message):
    print(f"[+] {message}")

def log_failure(message):
    print(f"[X] {message}")

def log_info(message):
    print(f"[!] {message}")

def create_template(data, filename):
    """Membuat template Excel dengan data yang diberikan"""
    if not os.path.exists(EXCEL_FOLDER):
        os.makedirs(EXCEL_FOLDER)
        log_success(f"Created directory: {EXCEL_FOLDER}")
    full_filepath = os.path.join(EXCEL_FOLDER, filename)
    df = pd.DataFrame(data)
    try:
        df.to_excel(full_filepath, index=False)
    except Exception as e:
        log_failure(f"Failed to save DataFrame to {full_filepath}: {e}")
        return
    try:
        wb = openpyxl.load_workbook(full_filepath)
    except Exception as e:
        log_failure(f"Error: Tidak dapat membuka file {full_filepath} - {e}")
        return
    ws = wb.active
    if ws is None:
        log_failure(f"Error: Tidak dapat mengakses worksheet dari file {full_filepath}")
        return
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = alignment
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = alignment
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max(15, min(max_length + 2, 50))
    try:
        wb.save(full_filepath)
        log_success(f"Template {full_filepath} berhasil dibuat!")
    except Exception as e:
        log_failure(f"Error saat menyimpan file {full_filepath}: {str(e)}")

def create_template():
    """Create template Excel file with pentest techniques"""
    # Create data directory if it doesn't exist
    os.makedirs('data/excel', exist_ok=True)
    
    # Template data
    template_data = {
        "techniques": [
            {
                "id": "1",
                "name": "Man-in-the-Middle (MITM) Attack",
                "category": "Network Security",
                "description": "Serangan yang memungkinkan penyerang untuk menyadap dan memodifikasi komunikasi antara dua pihak",
                "symptoms": [
                    "Traffic redirection",
                    "Data interception",
                    "Session hijacking"
                ],
                "solutions": [
                    "Gunakan enkripsi kuat pada access point (WPA2/WPA3)",
                    "Ganti kredensial login default pada router",
                    "Gunakan VPN untuk mengenkripsi lalu lintas jaringan",
                    "Implementasikan segmentasi jaringan (VLAN)",
                    "Aktifkan fitur isolasi klien pada access point"
                ],
                "tools": [
                    "Wireshark",
                    "Ettercap",
                    "SSLstrip",
                    "mitmproxy"
                ],
                "references": [
                    "https://www.cisco.com/c/en/us/solutions/security/man-in-the-middle-attack.html",
                    "https://www.owasp.org/index.php/Man-in-the-middle_attack"
                ],
                "priority": "Kritis",
                "problems": [
                    "Vulnerability pada komunikasi jaringan yang tidak terenkripsi",
                    "Lack of certificate validation",
                    "Insufficient traffic monitoring"
                ],
                "attack_signs": [
                    "Traffic redirection",
                    "Data interception",
                    "Session hijacking",
                    "Certificate errors"
                ]
            },
            {
                "id": "2",
                "name": "SQL Injection",
                "category": "Web Security",
                "description": "Serangan yang memanfaatkan kerentanan pada query database untuk mengakses atau memodifikasi data",
                "symptoms": [
                    "Data leakage",
                    "Unauthorized access",
                    "Database errors"
                ],
                "solutions": [
                    "Gunakan prepared statements",
                    "Implementasikan input validation",
                    "Terapkan principle of least privilege",
                    "Gunakan WAF (Web Application Firewall)",
                    "Lakukan regular security testing"
                ],
                "tools": [
                    "SQLmap",
                    "Burp Suite",
                    "OWASP ZAP",
                    "Acunetix"
                ],
                "references": [
                    "https://owasp.org/www-community/attacks/SQL_Injection",
                    "https://portswigger.net/web-security/sql-injection"
                ],
                "priority": "Kritis",
                "problems": [
                    "Unsanitized user input",
                    "Direct SQL query construction",
                    "Excessive database privileges"
                ],
                "attack_signs": [
                    "Database error messages",
                    "Unexpected data exposure",
                    "Abnormal database behavior"
                ]
            }
        ]
    }
    
    # Convert to DataFrame
    df = pd.DataFrame(template_data["techniques"])
    
    # Save as Excel
    df.to_excel(EXCEL_TEMPLATE, index=False)
    print(f"Template created at {EXCEL_TEMPLATE}")
    
    # Save as JSON
    with open(DATA_FILE, 'w') as f:
        json.dump(template_data, f, indent=2)
    print(f"JSON data saved at {DATA_FILE}")

def convert_excel_to_json():
    """Convert Excel template to JSON format"""
    try:
        # Read Excel file
        df = pd.read_excel(EXCEL_TEMPLATE)
        
        # Convert to dictionary
        data = {"techniques": df.to_dict('records')}
        
        # Save as JSON
        with open(DATA_FILE, 'w') as f:
            json.dump(data, f, indent=2)
        print(f"Data converted and saved to {DATA_FILE}")
        
    except Exception as e:
        print(f"Error converting Excel to JSON: {e}")
        raise

def create_all_templates():
    """Create all necessary templates"""
    try:
        template_path = create_template()
        print(f"Template created at: {template_path}")
        return True
    except Exception as e:
        print(f"Error creating templates: {e}")
        return False 